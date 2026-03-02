from __future__ import annotations

import os
import uuid
import shutil
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from typing import List, Optional, Any

from openpyxl import load_workbook

warnings.filterwarnings("ignore", message="Unknown extension is not supported*")
warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported*")
warnings.filterwarnings("ignore", message="Data Validation extension is not supported*")


@dataclass
class Demand:
    _id: str
    linha_excel: int

    urgente: Optional[str]
    status: Optional[str]
    prioridade: Optional[str]
    data_entrada: Optional[date]

    prazos_str: str
    data_entrega: Optional[date]

    projeto: Optional[str]
    descricao: Optional[str]
    comentario: Optional[str]

    id_azure: Optional[str]
    perc_conclusao_value: Optional[float]  # <-- agora é número (0..1)
    responsavel: Optional[str]

    reportar: bool
    nome: Optional[str]
    time_funcao: Optional[str]

    timing: str


# --------------------------
# Helpers
# --------------------------

def _to_date(v: Any) -> Optional[date]:
    """Converte valor do Excel para date (sem hora). Aceita yyyy-mm-dd e dd/mm/yyyy."""
    if v is None or v == "":
        return None

    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, date):
        return v

    s = str(v).strip()

    # tenta ISO primeiro: yyyy-mm-dd
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        pass

    # fallback BR: dd/mm/yyyy
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        return None


def _normalize_separators(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", ",")
    s = s.replace(";", ",")
    return s


def _parse_prazos(value: Any) -> List[date]:
    """Aceita date/datetime e strings com múltiplas datas (ISO ou BR), separadas por vírgula/; / quebra de linha."""
    if value is None or value == "":
        return []

    # Se já for data, vira lista com 1 item
    d0 = _to_date(value)
    if d0 and not isinstance(value, str):
        return [d0]

    s = _normalize_separators(str(value))
    parts = [p.strip() for p in s.split(",") if p.strip()]

    out: List[date] = []
    for p in parts:
        d = _to_date(p)
        if d:
            out.append(d)

    return sorted(set(out))


def _to_float_percent(v: Any) -> Optional[float]:
    """
    Converte % conclusão para float 0..1.
    Aceita:
    - número (0.75)
    - '75%' / '0,75' / '0.75'
    - vazio => None
    """
    if v is None or v == "":
        return None

    # se já for número
    if isinstance(v, (int, float)):
        f = float(v)
        # se vier como 75 (por engano), normaliza
        if f > 1.0:
            # se for 25, 50, 75, 100
            if f in (0, 25, 50, 75, 100):
                return f / 100.0
        return f

    s = str(v).strip().lower()
    s = s.replace(" ", "").replace("%", "")

    # decimal com vírgula
    s = s.replace(",", ".")

    try:
        f = float(s)
        if f > 1.0 and f in (0, 25, 50, 75, 100):
            f = f / 100.0
        return f
    except Exception:
        return None


def _calc_timing(status: Optional[str], prazos: List[date], entrega: Optional[date], hoje: date) -> str:
    if not status:
        return ""

    st = status.strip().lower()

    if st == "cancelado":
        return "Cancelado"

    if not prazos:
        return "Sem Prazo Definido"

    if st not in ("concluído", "concluido", "cancelado"):
        if min(prazos) < hoje and entrega is None:
            return "Em Atraso"
        return "Dentro do Prazo"

    if entrega is None:
        return "Concluído"

    p = min(prazos)
    if entrega > p:
        return "Concluída com Atraso"
    if entrega == p:
        return "Concluída no Prazo"
    return "Concluída antes do Prazo"


# --------------------------
# Store
# --------------------------

class ExcelStore:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.wb = None
        self.ws = None
        self.COL = {}
        self.demands: List[Demand] = []
        self.reload()

    def reload(self):
        self.wb = load_workbook(self.xlsx_path)
        if "Demandas" not in self.wb.sheetnames:
            raise ValueError("Aba 'Demandas' não encontrada no arquivo de dados.")
        self.ws = self.wb["Demandas"]
        self._map_columns()
        self._load_demands()

    def _map_columns(self):
        headers = {}
        for col in range(1, self.ws.max_column + 1):
            v = self.ws.cell(row=3, column=col).value
            if v:
                headers[str(v).strip().lower()] = col

        def col(name: str) -> Optional[int]:
            return headers.get(name.strip().lower())

        self.COL = {
            "urgente": col("urgente"),
            "status": col("status"),
            "prioridade": col("prioridade"),
            "data_entrada": col("data entrada"),
            "prazo": col("prazo"),
            "data_entrega": col("data entrega"),
            "projeto": col("projeto"),
            "descricao": col("descrição") or col("descricao"),
            "comentario": col("comentário") or col("comentario"),
            "id_azure": col("id azure"),
            "perc": col("% conclusão") or col("% conclusao"),
            "responsavel": col("responsável") or col("responsavel"),
            "reportar": col("reportar?"),
            "nome": col("nome"),
            "time_funcao": col("time/função") or col("time/funcao"),
            "_id": col("_id"),
        }

        if self.COL["_id"] is None:
            new_col = self.ws.max_column + 1
            self.ws.cell(row=3, column=new_col).value = "_id"
            try:
                self.ws.column_dimensions[self.ws.cell(row=3, column=new_col).column_letter].hidden = True
            except Exception:
                pass
            self.COL["_id"] = new_col
            self._save()

    def _cell(self, row: int, key: str):
        c = self.COL.get(key)
        if not c:
            return None
        return self.ws.cell(row=row, column=c)

    def _val(self, row: int, key: str):
        cell = self._cell(row, key)
        return cell.value if cell else None

    def _load_demands(self):
        self.demands = []
        hoje = date.today()

        for row in range(4, self.ws.max_row + 1):
            if not any(self.ws.cell(row=row, column=c).value not in (None, "") for c in range(1, self.ws.max_column + 1)):
                continue

            _id_cell = self._cell(row, "_id")
            _id = _id_cell.value if _id_cell else None
            if not _id:
                _id = str(uuid.uuid4())
                if _id_cell:
                    _id_cell.value = _id

            prazo_raw = self._val(row, "prazo")
            prazos = _parse_prazos(prazo_raw)

            entrega = _to_date(self._val(row, "data_entrega"))
            status = self._val(row, "status")

            timing = _calc_timing(status, prazos, entrega, hoje)

            prazos_str = str(prazo_raw) if prazo_raw is not None else ""

            perc_val = _to_float_percent(self._val(row, "perc"))

            d = Demand(
                _id=str(_id),
                linha_excel=row,
                urgente=self._val(row, "urgente"),
                status=status,
                prioridade=self._val(row, "prioridade"),
                data_entrada=_to_date(self._val(row, "data_entrada")),
                prazos_str=prazos_str,
                data_entrega=entrega,
                projeto=self._val(row, "projeto"),
                descricao=self._val(row, "descricao"),
                comentario=self._val(row, "comentario"),
                id_azure=self._val(row, "id_azure"),
                perc_conclusao_value=perc_val,
                responsavel=self._val(row, "responsavel"),
                reportar=str(self._val(row, "reportar")).strip().lower() == "sim",
                nome=self._val(row, "nome"),
                time_funcao=self._val(row, "time_funcao"),
                timing=timing,
            )
            self.demands.append(d)

        self._save()

    # ---------- Consultas ----------

    def get_by_id(self, _id: str) -> Optional[Demand]:
        for d in self.demands:
            if d._id == _id:
                return d
        return None

    def filter_by_prazo_date(self, d: date) -> List[Demand]:
        out: List[Demand] = []
        for x in self.demands:
            pr = _parse_prazos(x.prazos_str)
            if d in pr:
                out.append(x)
        return out

    def filter_overdue(self, hoje: date) -> List[Demand]:
        out: List[Demand] = []
        for x in self.demands:
            st = (x.status or "").strip()
            if st in ("Concluído", "Cancelado"):
                continue
            pr = _parse_prazos(x.prazos_str)
            if pr and min(pr) < hoje and x.data_entrega is None:
                out.append(x)
        return out

    def filter_by_entrega_date(self, d: date) -> List[Demand]:
        return [x for x in self.demands if x.data_entrega == d]

    def filter_pending(self) -> List[Demand]:
        return [x for x in self.demands if (x.status or "").strip() not in ("Concluído", "Cancelado")]

    def filter_concluido_between(self, start: date, end: date) -> List[Demand]:
        out: List[Demand] = []
        for x in self.demands:
            if (x.status or "").strip() != "Concluído":
                continue
            if x.data_entrega and start <= x.data_entrega <= end:
                out.append(x)
        return out

    # ---------- Escrita ----------

    def _backup(self):
        base = os.path.dirname(self.xlsx_path)
        bdir = os.path.join(base, "backups")
        os.makedirs(bdir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = os.path.join(bdir, f"Renato2_work_{ts}.xlsx")
        shutil.copy2(self.xlsx_path, dst)

    def _save(self):
        self.wb.save(self.xlsx_path)

    def add_demand(self, data: dict):
        self._backup()
        row = self.ws.max_row + 1
        self._cell(row, "_id").value = str(uuid.uuid4())
        self._write_row(row, data)
        self._save()
        self.reload()

    def update_demand(self, _id: str, data: dict):
        self._backup()
        for d in self.demands:
            if d._id == _id:
                self._write_row(d.linha_excel, data)
                self._save()
                self.reload()
                return
        raise ValueError("Demanda não encontrada para atualização.")

    def _write_row(self, row: int, data: dict):
        def w(key: str, value):
            c = self.COL.get(key)
            if c:
                self.ws.cell(row=row, column=c).value = value

        w("urgente", data.get("urgente"))
        w("status", data.get("status"))
        w("prioridade", data.get("prioridade"))
        w("data_entrada", data.get("data_entrada"))              # date
        w("prazo", data.get("prazos_str"))                       # string ISO
        w("data_entrega", data.get("data_entrega"))              # date|None
        w("projeto", data.get("projeto"))
        w("descricao", data.get("descricao"))
        w("comentario", data.get("comentario"))
        w("id_azure", data.get("id_azure"))
        w("perc", data.get("perc_conclusao_value"))              # float 0..1
        w("responsavel", data.get("responsavel"))
        w("reportar", "Sim" if data.get("reportar") else "Não")
        w("nome", data.get("nome"))
        w("time_funcao", data.get("time_funcao"))
